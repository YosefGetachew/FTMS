import json
from collections import defaultdict

import openpyxl


WORKBOOK = r"D:\OneDrive\Fle\weight loss data.xlsx"


def as_number(value):
    return value if isinstance(value, (int, float)) else None


wb_formula = openpyxl.load_workbook(WORKBOOK, data_only=False)
wb_values = openpyxl.load_workbook(WORKBOOK, data_only=True)
ws_formula = wb_formula.active
ws_values = wb_values[ws_formula.title]

headers = [ws_formula.cell(1, col).value for col in range(1, ws_formula.max_column + 1)]

formulas = []
negative_values = []
numeric_summary = defaultdict(lambda: {"count": 0, "min": None, "max": None, "negatives": 0})
formula_comparison_rows = []

for row in range(1, ws_formula.max_row + 1):
    for col in range(1, ws_formula.max_column + 1):
        formula_value = ws_formula.cell(row, col).value
        computed_value = ws_values.cell(row, col).value
        header = headers[col - 1] if row > 1 and col <= len(headers) else None

        if isinstance(formula_value, str) and formula_value.startswith("="):
            formulas.append(
                {
                    "cell": ws_formula.cell(row, col).coordinate,
                    "header": header,
                    "formula": formula_value,
                    "computed": computed_value,
                }
            )

        number = as_number(computed_value)
        if number is not None:
            stat = numeric_summary[header or f"Column {col}"]
            stat["count"] += 1
            stat["min"] = number if stat["min"] is None else min(stat["min"], number)
            stat["max"] = number if stat["max"] is None else max(stat["max"], number)
            if number < 0:
                stat["negatives"] += 1
                negative_values.append(
                    {
                        "cell": ws_values.cell(row, col).coordinate,
                        "row": row,
                        "header": header,
                        "value": number,
                        "row_values": [
                            ws_values.cell(row, c).value
                            for c in range(1, ws_values.max_column + 1)
                        ],
                        "row_formulas": [
                            ws_formula.cell(row, c).value
                            for c in range(1, ws_formula.max_column + 1)
                        ],
                    }
                )

for row in range(2, ws_values.max_row + 1):
    genotype, rep, nd, nu, wd, wu, current_result, sign = [
        ws_values.cell(row, col).value for col in range(1, 9)
    ]
    if not all(isinstance(value, (int, float)) for value in [nd, nu, wd, wu]):
        continue

    corrected = None
    if wu and (nd + nu):
        corrected = ((wu * nd) - (wd * nu)) / (wu * (nd + nu)) * 100

    avg_damaged = wd / nd if nd else None
    avg_undamaged = wu / nu if nu else None

    formula_comparison_rows.append(
        {
            "row": row,
            "genotype": genotype,
            "rep": rep,
            "ND": nd,
            "NU": nu,
            "WD": wd,
            "WU": wu,
            "current": current_result,
            "corrected_formula": corrected,
            "avg_damaged_weight": avg_damaged,
            "avg_undamaged_weight": avg_undamaged,
            "damaged_to_undamaged_avg_ratio": (
                avg_damaged / avg_undamaged
                if avg_damaged is not None and avg_undamaged
                else None
            ),
        }
    )

negative_comparisons = [
    row for row in formula_comparison_rows
    if isinstance(row["current"], (int, float)) and row["current"] < 0
]

print(json.dumps(
    {
        "sheets": [
            {
                "name": ws.title,
                "max_row": ws.max_row,
                "max_col": ws.max_column,
            }
            for ws in wb_formula.worksheets
        ],
        "headers": headers,
        "sample_rows": [
            [
                ws_values.cell(row, col).value
                for col in range(1, ws_values.max_column + 1)
            ]
            for row in range(1, min(ws_values.max_row, 12) + 1)
        ],
        "formulas_count": len(formulas),
        "formulas_sample": formulas[:50],
        "negative_count": len(negative_values),
        "negative_sample": negative_values[:80],
        "numeric_summary": dict(numeric_summary),
        "formula_comparison": {
            "current_range": [
                min(row["current"] for row in formula_comparison_rows if isinstance(row["current"], (int, float))),
                max(row["current"] for row in formula_comparison_rows if isinstance(row["current"], (int, float))),
            ],
            "corrected_formula_range": [
                min(row["corrected_formula"] for row in formula_comparison_rows if row["corrected_formula"] is not None),
                max(row["corrected_formula"] for row in formula_comparison_rows if row["corrected_formula"] is not None),
            ],
            "negative_rows_by_current_formula": len(negative_comparisons),
            "negative_rows_still_negative_with_corrected_formula": sum(
                1 for row in negative_comparisons
                if row["corrected_formula"] is not None and row["corrected_formula"] < 0
            ),
            "largest_negative_current_results": sorted(
                negative_comparisons,
                key=lambda item: item["current"],
            )[:20],
        },
    },
    indent=2,
    default=str,
))
